import pandas as pd
import openpyxl
from matplotlib import pyplot as plt
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment
import copy
import re


def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)  # 复制字体样式
        target_cell.border = copy.copy(source_cell.border)  # 复制边框样式
        target_cell.fill = copy.copy(source_cell.fill)  # 复制填充样式
        target_cell.number_format = source_cell.number_format  # 复制数字格式
        target_cell.protection = copy.copy(source_cell.protection)  # 复制保护属性
        target_cell.alignment = copy.copy(source_cell.alignment)  # 复制对齐方式


def copy_merged_cells(ws, start_col, end_col, offset):
    """复制合并单元格"""
    merged_cells_ranges = list(ws.merged_cells.ranges)  # 获取所有合并单元格范围
    for merged_cell in merged_cells_ranges:
        min_col, min_row, max_col, max_row = merged_cell.min_col, merged_cell.min_row, merged_cell.max_col, merged_cell.max_row
        if min_col >= start_col and max_col <= end_col:
            # 如果合并单元格在指定列范围内，复制并调整列偏移
            ws.merge_cells(start_row=min_row, start_column=min_col + offset, end_row=max_row,
                           end_column=max_col + offset)


def update_formula_references(formula, col_offset, original_col, target_col):
    """更新公式中的列引用"""

    def replace_col(match):
        col = match.group(1)
        row = match.group(2)
        col_index = column_index_from_string(col)
        new_col = get_column_letter(col_index + col_offset)
        return new_col + row

    def replace_absolute_col(match):
        col = match.group(1)
        row = match.group(2)
        if col == original_col:
            return f"${target_col}${row}"
        return f"${col}${row}"

    # 使用正则表达式匹配并替换公式中的列引用
    formula = re.sub(r'\$([A-Z]+)\$(\d+)', replace_absolute_col, formula)
    formula = re.sub(r'([A-Z]+)(\d+)', replace_col, formula)
    return formula


def copy_formula_and_style(ws, source_col, target_col, start_row, end_row, col_offset):
    """复制单元格内容、公式和样式"""
    original_col_letter = get_column_letter(source_col)
    target_col_letter = get_column_letter(target_col)

    for row in range(start_row, end_row + 1):
        for col_offset_inner in range(3):  # 复制3列
            source_cell = ws.cell(row=row, column=source_col + col_offset_inner)
            target_cell = ws.cell(row=row, column=target_col + col_offset_inner)
            if source_cell.data_type == 'f':  # 如果是公式，复制并更新公式引用
                updated_formula = update_formula_references(source_cell.value[1:], col_offset, original_col_letter,
                                                            target_col_letter)
                updated_formula = updated_formula.replace(f'{original_col_letter}$14', f'{target_col_letter}$14')
                target_cell.value = f"={updated_formula}"
            else:
                target_cell.value = source_cell.value  # 复制单元格值
            copy_cell_style(source_cell, target_cell)  # 复制单元格样式


def write_data(ws, data, start_col, col_offset):
    """将dataframe的值写入指定列"""
    df = pd.DataFrame(data)
    for key in data:
        row = int(key[1:])
        col_letter = get_column_letter(start_col + col_offset)
        ws[f'{col_letter}{row}'] = df.at[0, key]


def update(template_path, output_path, data1, data2, time_params):
    # 加载模板Excel文件
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    col = 0

    # 定义月份缩写
    month_map = {
        '01': 'Jan.',
        '02': 'Feb.',
        '03': 'Mar.',
        '04': 'Apr.',
        '05': 'May.',
        '06': 'Jun.',
        '07': 'Jul.',
        '08': 'Aug.',
        '09': 'Sep.',
        '10': 'Oct.',
        '11': 'Nov.',
        '12': 'Dec.'
    }

    # 更新B1单元格的月份缩写
    start_month = time_params[0]
    ws['B1'] = month_map[start_month[4:6]]
    ws['B1'].alignment = Alignment(horizontal='center')

    if len(time_params) > 1:
        for i, time in enumerate(time_params[1:], start=1):
            start_col = 2 + i * 3  # 计算新列的起始列号
            start_col_letter = get_column_letter(start_col)

            # 复制14Rx3C区域
            copy_formula_and_style(ws, 2, start_col, 2, 15, i * 3)  # 复制2到15行，3列宽度

            # 复制合并单元格
            copy_merged_cells(ws, 2, 4, i * 3)

            # 更新新列的月份缩写
            ws[f'{start_col_letter}1'] = month_map[time[4:6]]
            ws[f'{start_col_letter}1'].alignment = Alignment(horizontal='center')

            # 更新新列的绝对引用行
            for row in range(3, 14):  # 假设公式在3到13行
                cell = ws[f'{start_col_letter}{row}']
                if cell.data_type == 'f':
                    formula = cell.value
                    new_formula = formula.replace(f'B$14', f'{start_col_letter}$14')
                    cell.value = new_formula

            col = start_col

    # 将data1的值写入初始区域
    write_data(ws, data1, 2, 0)

    # 将data2的值写入新复制的区域
    write_data(ws, data2, col, 0)

    # 保存新的Excel文件
    wb.save(output_path)


# # Adjust categories to include subcategories under Labor and WWER separately
# categories = [
#     'Auto. Billing', 'Labor', 'Labor - BMS', 'Labor - CATS', 'WWER', 'WWER - BMS', 'WWER - CATS',
#     'Manual Billing', 'Manual - RBA BMS', 'Manual - ISB', 'Manual - SAP'
# ]
#
# # Values for April and May
# apr_values = [10000, 3000, 1000, 2000, 7000, 3000, 4000, 18000, 5000, 6000, 7000]
# may_values = [43320, 35555, 3232, 32323, 7765, 3232, 4533, 9174, 510, 3232, 5432]
#
# # Create the trend chart with clear color differentiation
# plt.figure(figsize=(14, 7))
# plt.plot(categories, apr_values, marker='o', label='April', color='blue', linestyle='-', linewidth=2)
# plt.plot(categories, may_values, marker='o', label='May', color='orange', linestyle='--', linewidth=2)
#
# # Set chart title and labels
# plt.title('Monthly Billing Structure (RMB)', fontsize=16)
# plt.xlabel('Categories', fontsize=12)
# plt.ylabel('Amount (RMB)', fontsize=12)
# plt.xticks(rotation=45, ha='right')
# plt.legend()
#
# # Save and display the chart
# image_path = "data/monthly_billing_structure_trend.png"
# plt.tight_layout()
# plt.savefig(image_path)
# plt.close()
#
# # Display the image
# display(Image(filename=image_path))

if __name__ == "__main__":
    template_path = 'data/testtest.xlsx'  # 替换为你的文件路径
    output_path = 'data/haha.xlsx'  # 替换为你的输出文件路径

    # 示例dataframe，替换为实际的dataframe
    data1 = {
        'B5': [1000],
        'B6': [2000],
        'B8': [3000],
        'B9': [4000],
        'B11': [5000],
        'B12': [6000],
        'B13': [7000]
    }
    data2 = {
        'B5': [3232],
        'B6': [32323],
        'B8': [3232],
        'B9': [4533],
        'B11': [510],
        'B12': [3232],
        'B13': [5432]
    }

    # 处理时间参数，根据需要复制14Rx3C区域
    time_params = ['202404', '202405']  # 替换为实际的时间参数

    update(template_path, output_path, data1, data2, time_params)
